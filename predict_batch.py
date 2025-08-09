import os
from PIL import Image
import numpy as np
from siamese_params import Siamese
import openpyxl
import time

if __name__ == "__main__":
    model = Siamese()
    image_1 = input('Input image_1 filename:')
    try:
        image_1 = Image.open(image_1)
    except:
        print('Image_1 Open Error! Try again!')
        exit(1)

    directory_path = r'../datasets/similarity_predict'  # 替换为你的文件夹路径
    probability_list = []  # 用于保存probability的列表

    start_time = time.time()

    for filename in os.listdir(directory_path):
        if filename.endswith(('.jpg', '.jpeg', '.png', '.bmp', '.gif')):  # 添加支持的图片格式
            image_2_path = os.path.join(directory_path, filename)
            try:
                image_2 = Image.open(image_2_path)
                probability = model.detect_image(image_1, image_2)
                probability_list.append((filename, probability))
            except:
                print(f'Error processing {filename}')
                continue

    end_time = time.time()
    total_time = end_time - start_time
    print(f'Total training time: {total_time} seconds')

    # 输出probability结果
    output_filename = r'..\output_file\result.xlsx'  # 指定要保存的XLSX文件名
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'time_count'

    # 写入标题行
    ws['A1'] = 'Image'
    ws['B1'] = 'Similarity'

    # 写入数据
    for i, (filename, probability) in enumerate(probability_list, start=2):
        cell_A = ws.cell(row=i, column=1)
        cell_B = ws.cell(row=i, column=2)
        cell_A.value = filename
        cell_B.value = probability.item()

    # 保存工作簿
    wb.save(output_filename)

    print(f'Results saved to {output_filename}')
